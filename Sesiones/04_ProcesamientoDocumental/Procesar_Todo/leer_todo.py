import os
import pymupdf
from docx import Document
from openpyxl import load_workbook


os.makedirs("textos", exist_ok=True)
def guardar_texto(texto:str, nombre_txt: str):
    with open(f"textos/{nombre_txt}", "w", encoding="utf-8") as f:
        f.write(texto)

def procesar_pdf(ruta, nombre_txt):
    pdf = pymupdf.open(ruta)
    texto = ""
    for pagina in pdf:
        texto += pagina.get_text() + "\n"
    nombre_txt = archivo.replace(".pdf", ".txt")

    guardar_texto(texto, nombre_txt)
    pdf.close()

def procesar_docs(ruta,nombre_txt):
    doc = Document(ruta)
    texto = ""
    for parrafo in doc.paragraphs:
        texto += parrafo.text + "\n"
    nombre_txt = archivo.replace(".docx", ".txt")
    guardar_texto(texto, nombre_txt)


def procesar_excel(ruta, nombre_txt):
    excel = load_workbook(ruta)
    texto = ""

    for hoja in excel:
        for fila in hoja.iter_rows(values_only=True):
            texto += " ".join(map(str, fila)) + "\n"

    nombre_txt = archivo.replace(".xlsx", ".txt")
    guardar_texto(texto, nombre_txt)


procesados = 0; no_soportados = 0; errores = 0

for archivo in os.listdir("Documentos"):
    try:
        if archivo.endswith(".pdf"):
            ruta = f"Documentos/{archivo}"
            procesar_pdf(ruta,archivo.replace(".pdf", ".txt"))
            procesados += 1

        elif archivo.endswith(".docx"):
            ruta = f"Documentos/{archivo}"
            procesar_docs(ruta, archivo.replace(".docx", ".txt"))
            procesados += 1


        elif archivo.endswith(".xlsx"):
            ruta = f"Documentos/{archivo}"
            procesar_excel(ruta,archivo.replace(".xlsx", ".txt"))
            procesados += 1

        else:
            print(f"Formato no soportado: {archivo}")
            no_soportados += 1

    except Exception as error:
        print(f"Error procesando {archivo}: {error}")
        errores += 1

print(f"Procesados correctamente: {procesados}")
print(f"Formatos no soportados: {no_soportados}")
print(f"Errores: {errores}")
