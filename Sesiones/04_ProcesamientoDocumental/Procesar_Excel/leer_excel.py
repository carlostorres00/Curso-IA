import os

from openpyxl import load_workbook

excel = load_workbook("Documentos/ejemplo.xlsx")

#print(excel.sheetnames)

#hoja = excel["NombreDeLaHoja"]
#hoja = excel.worksheets[0]
#hoja = excel.active #hoja activa es la que selecciona si abres el excel (no tiene porque ser la primera)

#print(hoja.title) #título de la hoja
#print(hoja["A1"].value) #valor de una celda

#for fila in hoja.iter_rows(values_only=True): #Recorro todas las filas y me lo devuelve como tupla
#    print(fila)


for archivo in os.listdir("Documentos"):
    print("Abriendo:", archivo)

    excel = load_workbook(f"Documentos/{archivo}")
    print("Excel abierto")

    texto = ""

    for hoja in excel:
        print("Procesando hoja:", hoja.title)

        for fila in hoja.iter_rows(values_only=True):
            texto += " ".join(map(str, fila)) + "\n"

    print("Texto generado")

    nombre_txt = archivo.replace(".xlsx", ".txt")

    with open(f"textos/{nombre_txt}", "w", encoding="utf-8") as f:
        f.write(texto)

    print("Archivo guardado")


