from openpyxl import load_workbook


def procesar_excel(ruta_entrada, ruta_salida):
    """
    Extrae el texto de todas las hojas de un archivo .xlsx y lo guarda en un .txt

    Args:
        ruta_entrada: ruta completa al archivo .xlsx a procesar
        ruta_salida: ruta completa donde guardar el .txt resultante
    """
    excel = load_workbook(ruta_entrada)

    texto = ""
    for hoja in excel:
        for fila in hoja.iter_rows(values_only=True):
            texto += " ".join(map(str, fila)) + "\n"

    with open(ruta_salida, "w", encoding="utf-8") as f:
        f.write(texto)

    return texto


